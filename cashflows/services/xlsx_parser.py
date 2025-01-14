from openpyxl.reader.excel import load_workbook 
from django.core.files.uploadedfile import UploadedFile
from openpyxl.worksheet.worksheet import Worksheet
from django.db.models import QuerySet
from datetime import datetime
from decimal import ROUND_HALF_UP, Decimal
from ..models import (
    Cashflow, 
    CashflowType,
    CashflowStatus,
    CashflowCategory, 
    CashflowSubcategory,
)
from ..config import CASHFLOWS_NAMES


def load_cashflows_from_file(xlsx_file: UploadedFile) -> QuerySet: 
    """Загружает ДДС из банковской выписки.
    Возвращает загруженные ДДС"""
    wb = load_workbook(filename=xlsx_file, data_only=True)
    sheet = wb.worksheets[0]

    __delete_all_loaded_cashflows()

    if _is_alpha_physical_person(sheet):  
        loaded_cashflows = _load_physical_person_statement(sheet)

    elif _is_alpha_individual_entrepreneur(sheet): 
        loaded_cashflows =_load_individual_entrepreneur_statement(sheet)

    return loaded_cashflows


def _load_physical_person_statement(sheet: Worksheet) -> QuerySet: 
    """Парсит данные о ДСС из выписки Альфа Банка для физлица
    Возвращает загруженные ДДС"""
    loaded_cashflows = []

    min_row = 2
    for index, row in enumerate(sheet.iter_rows(min_row=min_row, values_only=True), min_row):
        created_at_str, amount_float, category, cashflow_type_str, comment = row[0], float(row[7]), row[10], row[12], row[13]

        day, month, year = map(int, created_at_str.split('.'))
        created_at = datetime(year, month, day)
        amount = _round_decimal(Decimal(amount_float))
        comment = None if not comment else comment

        cashflow_status = CashflowStatus.get_personal_status()
        if cashflow_type_str == 'Пополнение': 
            cashflow_type = CashflowType.get_income_type()
        elif cashflow_type_str == 'Списание': 
            cashflow_type = CashflowType.get_expense_type()
        else: 
            raise Exception
        cashflow_category, _ = CashflowCategory.objects.get_or_create(
            name = CASHFLOWS_NAMES.Categories.PERSONAL_FROM_STATEMENT,
            cashflow_type = cashflow_type,
        )
        cashflow_subcategory, _ = CashflowSubcategory.objects.get_or_create(
            name=category, 
            cashflow_category = cashflow_category,
        )

        new_cashflow = Cashflow(
            created_at=created_at, 
            amount=amount, 
            cashflow_status=cashflow_status,
            cashflow_type=cashflow_type,
            cashflow_category=cashflow_category, 
            cashflow_subcategory=cashflow_subcategory,
            comment=comment,
        )
        new_cashflow.save()
        loaded_cashflows.append(new_cashflow)

    return loaded_cashflows


def _load_individual_entrepreneur_statement(sheet: Worksheet) -> QuerySet: 
    """Парсит данные о ДСС из выписки Альфа Банка для ИП. 
    Возвращает загруженные ДДС"""
    loaded_cashflows = []

    min_row = 13
    for index, row in enumerate(sheet.iter_rows(min_row=min_row, values_only=True), min_row):
        created_at_str = row[0]
        debit, credit = row[2], row[3] 

        day, month, year = map(int, created_at_str.split('.'))
        created_at = datetime(year, month, day)

        cashflow_status = CashflowStatus.get_business_status()
        if debit: 
            cashflow_type = CashflowType.get_expense_type() 
            amount_column = 2
        elif credit: 
            cashflow_type = CashflowType.get_income_type() 
            amount_column = 3

        amount_float = float(row[amount_column])
        amount = _round_decimal(Decimal(amount_float))

        cashflow_category, _ = CashflowCategory.objects.get_or_create(
            name = CASHFLOWS_NAMES.Categories.BUSINESS_FROM_STATEMENT,
            cashflow_type = cashflow_type,
        )
        cashflow_subcategory, _ = CashflowSubcategory.objects.get_or_create(
            name = CASHFLOWS_NAMES.Subcategories.BUSINESS_FROM_STATEMENT, 
            cashflow_category = cashflow_category,
        )

        new_cashflow = Cashflow(
            created_at=created_at, 
            amount=amount, 
            cashflow_status=cashflow_status,
            cashflow_type=cashflow_type,
            cashflow_category=cashflow_category, 
            cashflow_subcategory=cashflow_subcategory,
            comment=None,
        )
        new_cashflow.save()

        loaded_cashflows.append(new_cashflow)
        
    return loaded_cashflows
            
        

            
def _is_alpha_physical_person(sheet: Worksheet): 
    if sheet[1][0].value == 'Дата операции': 
        return True 
    return False


def _is_alpha_individual_entrepreneur(sheet: Worksheet): 
    if sheet[1][0].value == 'Выписка по счёту': 
        return True 
    return False


def _round_decimal(number: Decimal) -> Decimal: 
    return number.quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)



def __delete_all_loaded_cashflows(): 
    try:
        cashflow_type = CashflowType.get_income_type()
        cashflow_category = CashflowCategory.objects.get(
            name = CASHFLOWS_NAMES.Categories.PERSONAL_FROM_STATEMENT,
            cashflow_type = cashflow_type,
        )
        cashflow_category.delete() 
    except: 
        pass

    try:
        cashflow_category = CashflowCategory.objects.get(
            name = CASHFLOWS_NAMES.Categories.BUSINESS_FROM_STATEMENT,
            cashflow_type = cashflow_type,
        )
        cashflow_category.delete()  
    except: 
        pass

    try:
        cashflow_type = CashflowType.get_expense_type()
        cashflow_category = CashflowCategory.objects.get(
            name = CASHFLOWS_NAMES.Categories.BUSINESS_FROM_STATEMENT,
            cashflow_type = cashflow_type,
        )
        cashflow_category.delete() 
    except: 
        pass

    try:
        cashflow_type = CashflowType.get_expense_type()
        cashflow_category = CashflowCategory.objects.get(
            name = CASHFLOWS_NAMES.Categories.PERSONAL_FROM_STATEMENT,
            cashflow_type = cashflow_type,
        )
        cashflow_category.delete() 
    except: 
        pass